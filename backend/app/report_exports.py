from __future__ import annotations

from datetime import UTC, date, datetime
from io import BytesIO
import html
from pathlib import Path
import re
from zoneinfo import ZoneInfo

import arabic_reshaper
from bidi.algorithm import get_display
from openpyxl import Workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import LongTable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from .models import (
    Currency,
    Expense,
    ExpenseType,
    HouseExchange,
    JournalEntry,
    Order,
    User,
    Wallet,
    WalletAdjustment,
)


ISTANBUL_TZ = ZoneInfo("Europe/Istanbul")

C = {
    "darkBg": "1E3A5F",
    "midBg": "2D5F8F",
    "accent": "4A90D9",
    "lightBg": "EBF3FB",
    "white": "FFFFFF",
    "black": "1A1A2E",
    "green": "1A7A3C",
    "red": "B71C1C",
    "gray": "6B7280",
    "border": "BDD7EE",
}


def argb(hex_color: str) -> str:
    return f"FF{hex_color}"


def fmt_report_datetime(value: datetime) -> str:
    if value.tzinfo is None:
        value = value.replace(tzinfo=UTC)
    local = value.astimezone(ISTANBUL_TZ)
    return local.strftime("%d-%m-%Y %H:%M")


def fmt_money(value: object, decimals: int = 2) -> str:
    return f"{float(value):,.{decimals}f}"


def fmt_currency_money(value: object, currency: Currency | None) -> str:
    return fmt_money(value, currency.decimals if currency else 4)


def money_number_format(currency: Currency | None) -> str:
    decimals = currency.decimals if currency else 4
    return "#,##0" if decimals == 0 else f"#,##0.{''.join('0' for _ in range(decimals))}"


def solid_fill(hex_color: str) -> PatternFill:
    return PatternFill(fill_type="solid", fgColor=argb(hex_color))


def all_borders(color: str = C["border"]) -> Border:
    side = Side(style="thin", color=argb(color))
    return Border(top=side, bottom=side, left=side, right=side)


def safe_filename(value: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]", "_", value)


def date_key(value: datetime) -> str:
    return value.date().isoformat()


def title_name_part(value: str | None) -> str:
    if not value:
        return ""
    return value[:1].upper() + value[1:]


def user_full_name(user: User) -> str:
    parts = [title_name_part(user.name)]
    surname = title_name_part(user.surname)
    if surname:
        parts.append(surname)
    return " ".join(parts)


_PDF_FONT: str | None = None


def pdf_font() -> str:
    global _PDF_FONT
    if _PDF_FONT:
        return _PDF_FONT
    for font_path in [
        "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
        "/System/Library/Fonts/GeezaPro.ttc",
    ]:
        if Path(font_path).exists():
            pdfmetrics.registerFont(TTFont("ReportUnicode", font_path))
            _PDF_FONT = "ReportUnicode"
            return _PDF_FONT
    _PDF_FONT = "Helvetica"
    return _PDF_FONT


def pdf_styles() -> dict[str, ParagraphStyle]:
    font_name = pdf_font()
    styles = getSampleStyleSheet()
    return {
        "title": ParagraphStyle(
            "ReportTitle",
            parent=styles["Title"],
            fontName=font_name,
            fontSize=16,
            leading=20,
            alignment=TA_CENTER,
            textColor=colors.HexColor(f"#{C['darkBg']}"),
            spaceAfter=6,
        ),
        "subtitle": ParagraphStyle(
            "ReportSubtitle",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=10,
            leading=13,
            alignment=TA_CENTER,
            textColor=colors.HexColor(f"#{C['gray']}"),
            spaceAfter=10,
        ),
        "section": ParagraphStyle(
            "ReportSection",
            parent=styles["Heading2"],
            fontName=font_name,
            fontSize=10,
            leading=12,
            textColor=colors.HexColor(f"#{C['darkBg']}"),
            spaceBefore=8,
            spaceAfter=5,
        ),
        "cell": ParagraphStyle(
            "ReportCell",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=7,
            leading=9,
        ),
        "cellRight": ParagraphStyle(
            "ReportCellRight",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=7,
            leading=9,
            alignment=TA_RIGHT,
        ),
        "small": ParagraphStyle(
            "ReportSmall",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=8,
            leading=10,
            textColor=colors.HexColor(f"#{C['gray']}"),
        ),
        "smallCenter": ParagraphStyle(
            "ReportSmallCenter",
            parent=styles["Normal"],
            fontName=font_name,
            fontSize=8,
            leading=10,
            alignment=TA_CENTER,
            textColor=colors.HexColor(f"#{C['gray']}"),
        ),
    }


ARABIC_SCRIPT_RE = re.compile(r"[\u0600-\u06FF\u0750-\u077F\u08A0-\u08FF]")


def shape_rtl_text(value: str) -> str:
    if not ARABIC_SCRIPT_RE.search(value):
        return value
    return "\n".join(get_display(arabic_reshaper.reshape(line)) for line in value.splitlines())


def pdf_paragraph(value: object, style: ParagraphStyle, color: str | None = None) -> Paragraph:
    text = html.escape(shape_rtl_text("" if value is None else str(value))).replace("\n", "<br/>")
    if color:
        text = f'<font color="#{color}">{text}</font>'
    return Paragraph(text, style)


def pdf_table(
    headers: list[str],
    rows: list[list[object]],
    widths: list[float],
    *,
    right_columns: set[int] | None = None,
    color_cells: dict[tuple[int, int], str] | None = None,
) -> LongTable:
    right_columns = right_columns or set()
    color_cells = color_cells or {}
    styles = pdf_styles()
    table_data: list[list[Paragraph]] = [
        [pdf_paragraph(header, styles["cell"], C["white"]) for header in headers]
    ]
    for row_index, row in enumerate(rows):
        pdf_row = []
        for col_index, value in enumerate(row):
            style = styles["cellRight"] if col_index in right_columns else styles["cell"]
            pdf_row.append(pdf_paragraph(value, style, color_cells.get((row_index, col_index))))
        table_data.append(pdf_row)

    table = LongTable(table_data, colWidths=widths, repeatRows=1, splitByRow=True)
    table_style: list[tuple] = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(f"#{C['darkBg']}")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), pdf_font()),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor(f"#{C['border']}")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    for index in range(len(rows)):
        if index % 2 == 1:
            table_style.append(("BACKGROUND", (0, index + 1), (-1, index + 1), colors.HexColor("#F8FAFC")))
    table.setStyle(TableStyle(table_style))
    return table


def build_client_statement_xlsx(
    *,
    user: User,
    wallets: list[Wallet],
    currencies: list[Currency],
    orders: list[Order],
    journals: list[JournalEntry],
    wallet_adjustments: list[WalletAdjustment],
    user_wallet_ids: set[int],
    from_date: date,
    to_date: date,
) -> tuple[bytes, str]:
    curr_map = {currency.ticker: currency for currency in currencies}
    full_name = user_full_name(user)
    period_label = f"{from_date.strftime('%d/%m/%Y')} — {to_date.strftime('%d/%m/%Y')}"
    generated_at = fmt_report_datetime(datetime.now(UTC))

    tx_rows: list[dict[str, object]] = []
    for order in orders:
        in_curr = curr_map.get(order.currency_in_id)
        out_curr = curr_map.get(order.currency_out_id)
        tx_rows.append(
            {
                "date": order.created_at,
                "dateLabel": fmt_report_datetime(order.created_at),
                "category": "FX Order",
                "type": order.order_type.value,
                "description": (
                    f"{in_curr.name if in_curr else order.currency_in_id} → "
                    f"{out_curr.name if out_curr else order.currency_out_id} "
                    f"@ {float(order.exchange_rate):.4f}"
                ),
                "sent": f"{fmt_currency_money(order.amount_in, in_curr)} {order.currency_in_id}",
                "received": f"{fmt_currency_money(order.amount_out, out_curr)} {order.currency_out_id}",
                "note": order.description or "",
                "status": "Voided" if order.voided_at else "Active",
                "voided": bool(order.voided_at),
            }
        )

    for entry in journals:
        is_out = entry.from_wallet_id in user_wallet_ids
        currency = curr_map.get(entry.currency_id)
        tx_rows.append(
            {
                "date": entry.created_at,
                "dateLabel": fmt_report_datetime(entry.created_at),
                "category": "Transfer",
                "type": "Sent" if is_out else "Received",
                "description": f"{currency.name if currency else entry.currency_id} transfer",
                "sent": f"{fmt_currency_money(entry.amount, currency)} {entry.currency_id}" if is_out else "",
                "received": "" if is_out else f"{fmt_currency_money(entry.amount, currency)} {entry.currency_id}",
                "note": entry.description or "",
                "status": "Voided" if entry.voided_at else "Active",
                "voided": bool(entry.voided_at),
            }
        )

    for adjustment in wallet_adjustments:
        currency = curr_map.get(adjustment.currency_id)
        amount = adjustment.amount_delta
        tx_rows.append(
            {
                "date": adjustment.created_at,
                "dateLabel": fmt_report_datetime(adjustment.created_at),
                "category": "Wallet Adjustment",
                "type": adjustment.currency_id,
                "description": f"{currency.name if currency else adjustment.currency_id} balance adjustment",
                "sent": f"{fmt_currency_money(abs(amount), currency)} {adjustment.currency_id}" if amount < 0 else "",
                "received": f"{fmt_currency_money(amount, currency)} {adjustment.currency_id}" if amount > 0 else "",
                "note": adjustment.reason,
                "status": "Active",
                "voided": False,
            }
        )

    tx_rows.sort(key=lambda item: item["date"])

    wb = Workbook()
    wb.creator = "FX Ledger"
    wb.created = datetime.now()
    ws = wb.active
    ws.title = "Client Statement"
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    widths = [2, 20, 12, 10, 28, 22, 22, 32, 10, 2]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width

    def style(
        cell_ref: str,
        *,
        bold: bool = False,
        size: int = 10,
        color: str = C["black"],
        fill: str | None = None,
        border: Border | None = None,
        horizontal: str = "left",
        vertical: str = "middle",
        wrap_text: bool = False,
        number_format: str | None = None,
    ) -> None:
        cell = ws[cell_ref]
        cell.font = Font(name="Calibri", bold=bold, size=size, color=argb(color))
        if fill:
            cell.fill = solid_fill(fill)
        if border:
            cell.border = border
        cell.alignment = Alignment(
            horizontal=horizontal,
            vertical="center" if vertical == "middle" else vertical,
            wrap_text=wrap_text,
        )
        if number_format:
            cell.number_format = number_format

    row = 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    ws[f"B{row}"] = "CLIENT STATEMENT"
    style(f"B{row}", bold=True, size=18, color=C["white"], fill=C["darkBg"], horizontal="center")
    ws.row_dimensions[row].height = 38
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    ws[f"B{row}"] = full_name
    style(f"B{row}", bold=True, size=13, color=C["white"], fill=C["darkBg"], horizontal="center")
    ws.row_dimensions[row].height = 24
    row += 1

    ws.row_dimensions[row].height = 6
    row += 1

    for label, value in [("Report Period", period_label), ("Generated", generated_at)]:
        ws[f"B{row}"] = label
        style(f"B{row}", bold=True, size=10, color=C["gray"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=9)
        ws[f"C{row}"] = value
        style(f"C{row}", size=10, color=C["black"])
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    ws[f"B{row}"] = "WALLET BALANCES"
    style(
        f"B{row}",
        bold=True,
        size=11,
        color=C["white"],
        fill=C["midBg"],
        border=all_borders(C["midBg"]),
    )
    ws.row_dimensions[row].height = 22
    row += 1

    for col, label, align in [("B", "Currency", "left"), ("C", "Balance", "right")]:
        ws[f"{col}{row}"] = label
        style(
            f"{col}{row}",
            bold=True,
            size=10,
            color=C["white"],
            fill=C["accent"],
            border=all_borders(),
            horizontal=align,
        )
    ws.row_dimensions[row].height = 18
    row += 1

    nonzero_wallets = sorted(
        [wallet for wallet in wallets if wallet.balance != 0],
        key=lambda wallet: wallet.currency_id,
    )

    if not nonzero_wallets:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
        ws[f"B{row}"] = "No non-zero wallets"
        style(f"B{row}", size=10, color=C["gray"])
        row += 1
    else:
        for index, wallet in enumerate(nonzero_wallets):
            fill = C["lightBg"] if index % 2 == 1 else C["white"]
            currency = curr_map.get(wallet.currency_id)
            balance = float(wallet.balance)
            ws[f"B{row}"] = currency.name if currency else wallet.currency_id
            ws[f"C{row}"] = abs(balance)
            style(f"B{row}", size=10, fill=fill, border=all_borders())
            style(
                f"C{row}",
                size=10,
                color=C["red"] if balance > 0 else C["green"],
                fill=fill,
                border=all_borders(),
                horizontal="right",
                number_format=money_number_format(currency),
            )
            ws.row_dimensions[row].height = 18
            row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    ws[f"B{row}"] = CellRichText(
        TextBlock(InlineFont(rFont="Calibri", sz=8, color=argb(C["green"])), "Green=Lintu owes client"),
        TextBlock(InlineFont(rFont="Calibri", sz=8, color=argb(C["black"])), "\n"),
        TextBlock(InlineFont(rFont="Calibri", sz=8, color=argb(C["red"])), "Red=client owes Lintu"),
    )
    ws[f"B{row}"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 20
    row += 1

    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    ws[f"B{row}"] = f"TRANSACTION HISTORY  ({len(tx_rows)} transactions)"
    style(
        f"B{row}",
        bold=True,
        size=11,
        color=C["white"],
        fill=C["midBg"],
        border=all_borders(C["midBg"]),
    )
    ws.row_dimensions[row].height = 22
    row += 1

    tx_headers = [
        ("B", "Date & Time", "left"),
        ("C", "Category", "center"),
        ("D", "Type", "center"),
        ("E", "Description", "left"),
        ("F", "Sent (Debit)", "right"),
        ("G", "Received (Credit)", "right"),
        ("H", "Note", "left"),
        ("I", "Status", "center"),
    ]
    for col, label, align in tx_headers:
        ws[f"{col}{row}"] = label
        style(
            f"{col}{row}",
            bold=True,
            size=10,
            color=C["white"],
            fill=C["darkBg"],
            border=all_borders(C["darkBg"]),
            horizontal=align,
        )
    ws.row_dimensions[row].height = 20
    row += 1

    if not tx_rows:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
        ws[f"B{row}"] = "No transactions in this period"
        style(f"B{row}", size=10, color=C["gray"], horizontal="center")
        ws.row_dimensions[row].height = 20
        row += 1
    else:
        for index, tx in enumerate(tx_rows):
            fill = "FFF3F3" if tx["voided"] else C["lightBg"] if index % 2 == 1 else C["white"]
            values = {
                "B": tx["dateLabel"],
                "C": tx["category"],
                "D": tx["type"],
                "E": tx["description"],
                "F": tx["sent"],
                "G": tx["received"],
                "H": tx["note"],
                "I": tx["status"],
            }
            for col, value in values.items():
                ws[f"{col}{row}"] = value
                is_right = col in {"F", "G"}
                is_center = col in {"C", "D", "I"}
                is_voided_status = col == "I" and bool(tx["voided"])
                text_color = C["black"]
                if is_voided_status:
                    text_color = C["red"]
                elif tx["voided"]:
                    text_color = C["gray"]
                elif col == "F" and tx["sent"]:
                    text_color = C["red"]
                elif col == "G" and tx["received"]:
                    text_color = C["green"]

                style(
                    f"{col}{row}",
                    bold=col in {"F", "G"},
                    size=9,
                    color=text_color,
                    fill=fill,
                    border=all_borders(),
                    horizontal="right" if is_right else "center" if is_center else "left",
                    wrap_text=col == "H",
                )
            ws.row_dimensions[row].height = 18
            row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=9)
    ws[f"B{row}"] = (
        f"This statement was generated on {generated_at} and reflects all transactions "
        "recorded in the system for the selected period."
    )
    style(f"B{row}", size=8, color=C["gray"], horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 28

    output = BytesIO()
    wb.save(output)
    filename = f"{safe_filename(full_name)}_statement_{from_date.isoformat()}_to_{to_date.isoformat()}.xlsx"
    return output.getvalue(), filename


def build_client_statement_pdf(
    *,
    user: User,
    wallets: list[Wallet],
    currencies: list[Currency],
    orders: list[Order],
    journals: list[JournalEntry],
    wallet_adjustments: list[WalletAdjustment],
    user_wallet_ids: set[int],
    from_date: date,
    to_date: date,
) -> tuple[bytes, str]:
    curr_map = {currency.ticker: currency for currency in currencies}
    full_name = user_full_name(user)
    period_label = f"{from_date.strftime('%d/%m/%Y')} — {to_date.strftime('%d/%m/%Y')}"
    generated_at = fmt_report_datetime(datetime.now(UTC))
    styles = pdf_styles()
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    story = [
        pdf_paragraph("CLIENT STATEMENT", styles["title"]),
        pdf_paragraph(full_name, styles["subtitle"]),
        pdf_paragraph(f"Report Period: {period_label}    Generated: {generated_at}", styles["small"]),
        Spacer(1, 4),
        pdf_paragraph("WALLET BALANCES", styles["section"]),
    ]

    wallet_rows: list[list[object]] = []
    wallet_colors: dict[tuple[int, int], str] = {}
    nonzero_wallets = sorted(
        [wallet for wallet in wallets if wallet.balance != 0],
        key=lambda wallet: wallet.currency_id,
    )
    for index, wallet in enumerate(nonzero_wallets):
        currency = curr_map.get(wallet.currency_id)
        balance = float(wallet.balance)
        wallet_rows.append([
            currency.name if currency else wallet.currency_id,
            fmt_currency_money(abs(wallet.balance), currency),
        ])
        wallet_colors[(index, 1)] = C["red"] if balance > 0 else C["green"]
    if not wallet_rows:
        wallet_rows.append(["No non-zero wallets", ""])
    story.append(pdf_table(["Currency", "Balance"], wallet_rows, [110 * mm, 50 * mm], right_columns={1}, color_cells=wallet_colors))
    story.append(pdf_paragraph("Green=Lintu owes client", styles["smallCenter"], C["green"]))
    story.append(pdf_paragraph("Red=client owes Lintu", styles["smallCenter"], C["red"]))
    story.append(Spacer(1, 6))

    tx_rows: list[dict[str, object]] = []
    for order in orders:
        in_curr = curr_map.get(order.currency_in_id)
        out_curr = curr_map.get(order.currency_out_id)
        tx_rows.append(
            {
                "date": order.created_at,
                "dateLabel": fmt_report_datetime(order.created_at),
                "category": "FX Order",
                "type": order.order_type.value,
                "description": (
                    f"{in_curr.name if in_curr else order.currency_in_id} → "
                    f"{out_curr.name if out_curr else order.currency_out_id} "
                    f"@ {float(order.exchange_rate):.4f}"
                ),
                "sent": f"{fmt_currency_money(order.amount_in, in_curr)} {order.currency_in_id}",
                "received": f"{fmt_currency_money(order.amount_out, out_curr)} {order.currency_out_id}",
                "note": order.description or "",
                "status": "Voided" if order.voided_at else "Active",
            }
        )

    for entry in journals:
        is_out = entry.from_wallet_id in user_wallet_ids
        currency = curr_map.get(entry.currency_id)
        tx_rows.append(
            {
                "date": entry.created_at,
                "dateLabel": fmt_report_datetime(entry.created_at),
                "category": "Transfer",
                "type": "Sent" if is_out else "Received",
                "description": f"{currency.name if currency else entry.currency_id} transfer",
                "sent": f"{fmt_currency_money(entry.amount, currency)} {entry.currency_id}" if is_out else "",
                "received": "" if is_out else f"{fmt_currency_money(entry.amount, currency)} {entry.currency_id}",
                "note": entry.description or "",
                "status": "Voided" if entry.voided_at else "Active",
            }
        )

    for adjustment in wallet_adjustments:
        currency = curr_map.get(adjustment.currency_id)
        amount = adjustment.amount_delta
        tx_rows.append(
            {
                "date": adjustment.created_at,
                "dateLabel": fmt_report_datetime(adjustment.created_at),
                "category": "Wallet Adjustment",
                "type": adjustment.currency_id,
                "description": f"{currency.name if currency else adjustment.currency_id} balance adjustment",
                "sent": f"{fmt_currency_money(abs(amount), currency)} {adjustment.currency_id}" if amount < 0 else "",
                "received": f"{fmt_currency_money(amount, currency)} {adjustment.currency_id}" if amount > 0 else "",
                "note": adjustment.reason,
                "status": "Active",
            }
        )
    tx_rows.sort(key=lambda item: item["date"])

    story.append(pdf_paragraph(f"TRANSACTION HISTORY ({len(tx_rows)} transactions)", styles["section"]))
    table_rows = [
        [
            row["dateLabel"],
            row["category"],
            row["type"],
            row["description"],
            row["sent"],
            row["received"],
            row["note"],
            row["status"],
        ]
        for row in tx_rows
    ] or [["No transactions in this period", "", "", "", "", "", "", ""]]
    story.append(
        pdf_table(
            ["Date & Time", "Category", "Type", "Description", "Sent (Debit)", "Received (Credit)", "Note", "Status"],
            table_rows,
            [28 * mm, 24 * mm, 18 * mm, 58 * mm, 34 * mm, 36 * mm, 55 * mm, 20 * mm],
            right_columns={4, 5},
        )
    )
    story.append(Spacer(1, 6))
    story.append(
        pdf_paragraph(
            f"This statement was generated on {generated_at} and reflects all transactions recorded in the system for the selected period.",
            styles["small"],
        )
    )
    doc.build(story)
    filename = f"{safe_filename(full_name)}_statement_{from_date.isoformat()}_to_{to_date.isoformat()}.pdf"
    return output.getvalue(), filename


def _format_period_label(from_date: date | None, to_date: date | None) -> str:
    if from_date and to_date:
        return f"{from_date.strftime('%d/%m/%Y')} — {to_date.strftime('%d/%m/%Y')}"
    if from_date:
        return f"From {from_date.strftime('%d/%m/%Y')}"
    if to_date:
        return f"Through {to_date.strftime('%d/%m/%Y')}"
    return "All time"


def build_full_activity_report_xlsx(
    *,
    users: list[User],
    currencies: list[Currency],
    wallets: list[Wallet],
    orders: list[Order],
    house_exchanges: list[HouseExchange],
    expenses: list[Expense],
    journals: list[JournalEntry],
    wallet_adjustments: list[WalletAdjustment],
    from_date: date | None,
    to_date: date | None,
) -> tuple[bytes, str]:
    generated_at = fmt_report_datetime(datetime.now(UTC))
    period_label = _format_period_label(from_date, to_date)
    user_map = {user.id: user for user in users}
    wallet_map = {wallet.id: wallet for wallet in wallets}
    currency_map = {currency.ticker: currency for currency in currencies}

    def user_label(user_id: int | None) -> str:
        if user_id is None:
            return ""
        user = user_map.get(user_id)
        if not user:
            return str(user_id)
        return user_full_name(user)

    def note_with_void(note: str | None, void_reason: str | None) -> str:
        parts = []
        if note:
            parts.append(note)
        if void_reason:
            parts.append(f"Void reason: {void_reason}")
        return " | ".join(parts)

    def wallet_user_label(wallet_id: int) -> str:
        wallet = wallet_map.get(wallet_id)
        if not wallet:
            return f"Wallet #{wallet_id}"
        return user_label(wallet.user_id)

    tx_rows: list[dict[str, object]] = []
    for order in orders:
        in_curr = currency_map.get(order.currency_in_id)
        out_curr = currency_map.get(order.currency_out_id)
        tx_rows.append(
            {
                "date": order.created_at,
                "dateLabel": fmt_report_datetime(order.created_at),
                "category": "FX Order",
                "party": user_label(order.client_id),
                "type": order.order_type.value,
                "description": (
                    f"{in_curr.name if in_curr else order.currency_in_id} → "
                    f"{out_curr.name if out_curr else order.currency_out_id} "
                    f"@ {float(order.exchange_rate):.4f}"
                ),
                "sent": f"{fmt_currency_money(order.amount_in, in_curr)} {order.currency_in_id}",
                "received": f"{fmt_currency_money(order.amount_out, out_curr)} {order.currency_out_id}",
                "note": note_with_void(order.description, order.void_reason),
                "status": "Voided" if order.voided_at else "Active",
                "voided": bool(order.voided_at),
            }
        )

    for exchange in house_exchanges:
        from_curr = currency_map.get(exchange.currency_from_id)
        to_curr = currency_map.get(exchange.currency_to_id)
        tx_rows.append(
            {
                "date": exchange.created_at,
                "dateLabel": fmt_report_datetime(exchange.created_at),
                "category": "House Exchange",
                "party": user_label(exchange.house_id),
                "type": "Exchange",
                "description": (
                    f"{from_curr.name if from_curr else exchange.currency_from_id} → "
                    f"{to_curr.name if to_curr else exchange.currency_to_id} "
                    f"@ {float(exchange.exchange_rate):.4f}"
                ),
                "sent": f"{fmt_currency_money(exchange.amount_from, from_curr)} {exchange.currency_from_id}",
                "received": f"{fmt_currency_money(exchange.amount_to, to_curr)} {exchange.currency_to_id}",
                "note": note_with_void(exchange.description, exchange.void_reason),
                "status": "Voided" if exchange.voided_at else "Active",
                "voided": bool(exchange.voided_at),
            }
        )

    for expense in expenses:
        currency = currency_map.get(expense.currency_id)
        kind_label = "Withdrawal" if expense.expense_type == ExpenseType.WITHDRAWAL else "Expense"
        party = user_label(expense.house_id)
        if expense.recipient_user_id is not None:
            party = f"{party} → {user_label(expense.recipient_user_id)}"
        tx_rows.append(
            {
                "date": expense.created_at,
                "dateLabel": fmt_report_datetime(expense.created_at),
                "category": kind_label,
                "party": party,
                "type": expense.currency_id,
                "description": f"{currency.name if currency else expense.currency_id} {kind_label.lower()}",
                "sent": f"{fmt_currency_money(expense.amount, currency)} {expense.currency_id}",
                "received": "",
                "note": note_with_void(expense.description, expense.void_reason),
                "status": "Voided" if expense.voided_at else "Active",
                "voided": bool(expense.voided_at),
            }
        )

    for entry in journals:
        currency = currency_map.get(entry.currency_id)
        tx_rows.append(
            {
                "date": entry.created_at,
                "dateLabel": fmt_report_datetime(entry.created_at),
                "category": "Transfer",
                "party": f"{wallet_user_label(entry.from_wallet_id)} → {wallet_user_label(entry.to_wallet_id)}",
                "type": entry.currency_id,
                "description": f"{currency.name if currency else entry.currency_id} transfer",
                "sent": f"{fmt_currency_money(entry.amount, currency)} {entry.currency_id}",
                "received": f"{fmt_currency_money(entry.amount, currency)} {entry.currency_id}",
                "note": note_with_void(entry.description, entry.void_reason),
                "status": "Voided" if entry.voided_at else "Active",
                "voided": bool(entry.voided_at),
            }
        )

    for adjustment in wallet_adjustments:
        wallet = wallet_map.get(adjustment.wallet_id)
        currency = currency_map.get(adjustment.currency_id)
        amount = adjustment.amount_delta
        tx_rows.append(
            {
                "date": adjustment.created_at,
                "dateLabel": fmt_report_datetime(adjustment.created_at),
                "category": "Wallet Adjustment",
                "party": user_label(wallet.user_id) if wallet else f"Wallet #{adjustment.wallet_id}",
                "type": adjustment.currency_id,
                "description": f"{currency.name if currency else adjustment.currency_id} balance adjustment",
                "sent": f"{fmt_currency_money(abs(amount), currency)} {adjustment.currency_id}" if amount < 0 else "",
                "received": f"{fmt_currency_money(amount, currency)} {adjustment.currency_id}" if amount > 0 else "",
                "note": adjustment.reason,
                "status": "Active",
                "voided": False,
            }
        )

    tx_rows.sort(key=lambda item: item["date"])

    wb = Workbook()
    wb.creator = "FX Ledger"
    wb.created = datetime.now()
    _ = wallet_adjustments
    ws = wb.active
    ws.title = "Activity Report"
    ws.sheet_view.showGridLines = False
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1

    widths = [2, 20, 16, 30, 12, 30, 22, 22, 32, 10, 2]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + index)].width = width

    def style(
        cell_ref: str,
        *,
        bold: bool = False,
        size: int = 10,
        color: str = C["black"],
        fill: str | None = None,
        border: Border | None = None,
        horizontal: str = "left",
        wrap_text: bool = False,
    ) -> None:
        cell = ws[cell_ref]
        cell.font = Font(name="Calibri", bold=bold, size=size, color=argb(color))
        if fill:
            cell.fill = solid_fill(fill)
        if border:
            cell.border = border
        cell.alignment = Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap_text)

    row = 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws[f"B{row}"] = "FULL ACTIVITY REPORT"
    style(f"B{row}", bold=True, size=18, color=C["white"], fill=C["darkBg"], horizontal="center")
    ws.row_dimensions[row].height = 38
    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws[f"B{row}"] = "FX Orders & Transfers"
    style(f"B{row}", bold=True, size=13, color=C["white"], fill=C["darkBg"], horizontal="center")
    ws.row_dimensions[row].height = 24
    row += 1

    ws.row_dimensions[row].height = 6
    row += 1

    for label, value in [
        ("Report Period", period_label),
        ("Generated", generated_at),
    ]:
        ws[f"B{row}"] = label
        style(f"B{row}", bold=True, size=10, color=C["gray"])
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=10)
        ws[f"C{row}"] = value
        style(f"C{row}", size=10, color=C["black"])
        ws.row_dimensions[row].height = 18
        row += 1

    row += 1

    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws[f"B{row}"] = f"TRANSACTION HISTORY  ({len(tx_rows)} transactions)"
    style(
        f"B{row}",
        bold=True,
        size=11,
        color=C["white"],
        fill=C["midBg"],
        border=all_borders(C["midBg"]),
    )
    ws.row_dimensions[row].height = 22
    row += 1

    headers = [
        ("B", "Date & Time", "left"),
        ("C", "Category", "center"),
        ("D", "Party", "left"),
        ("E", "Type", "center"),
        ("F", "Description", "left"),
        ("G", "Sent (Debit)", "right"),
        ("H", "Received (Credit)", "right"),
        ("I", "Note", "left"),
        ("J", "Status", "center"),
    ]
    for col, label, align in headers:
        ws[f"{col}{row}"] = label
        style(
            f"{col}{row}",
            bold=True,
            size=10,
            color=C["white"],
            fill=C["darkBg"],
            border=all_borders(C["darkBg"]),
            horizontal=align,
        )
    ws.row_dimensions[row].height = 20
    row += 1

    if not tx_rows:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        ws[f"B{row}"] = "No transactions in this period"
        style(f"B{row}", size=10, color=C["gray"], horizontal="center")
        ws.row_dimensions[row].height = 20
        row += 1
    else:
        for index, tx in enumerate(tx_rows):
            fill = "FFF3F3" if tx["voided"] else C["lightBg"] if index % 2 == 1 else C["white"]
            values = {
                "B": tx["dateLabel"],
                "C": tx["category"],
                "D": tx["party"],
                "E": tx["type"],
                "F": tx["description"],
                "G": tx["sent"],
                "H": tx["received"],
                "I": tx["note"],
                "J": tx["status"],
            }
            for col, value in values.items():
                ws[f"{col}{row}"] = value
                is_right = col in {"G", "H"}
                is_center = col in {"C", "E", "J"}
                is_voided_status = col == "J" and bool(tx["voided"])
                text_color = C["black"]
                if is_voided_status:
                    text_color = C["red"]
                elif tx["voided"]:
                    text_color = C["gray"]
                elif col == "G" and tx["sent"]:
                    text_color = C["red"]
                elif col == "H" and tx["received"]:
                    text_color = C["green"]

                style(
                    f"{col}{row}",
                    bold=col in {"G", "H"},
                    size=9,
                    color=text_color,
                    fill=fill,
                    border=all_borders(),
                    horizontal="right" if is_right else "center" if is_center else "left",
                    wrap_text=col in {"D", "F", "I"},
                )
            ws.row_dimensions[row].height = 18
            row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws[f"B{row}"] = "NON-ZERO WALLETS"
    style(
        f"B{row}",
        bold=True,
        size=11,
        color=C["white"],
        fill=C["midBg"],
        border=all_borders(C["midBg"]),
    )
    ws.row_dimensions[row].height = 22
    row += 1

    wallet_headers = [
        ("B", "Name", "left"),
        ("C", "Role", "center"),
        ("D", "Currency", "left"),
        ("E", "Balance", "right"),
    ]
    for col, label, align in wallet_headers:
        ws[f"{col}{row}"] = label
        style(
            f"{col}{row}",
            bold=True,
            size=10,
            color=C["white"],
            fill=C["darkBg"],
            border=all_borders(C["darkBg"]),
            horizontal=align,
        )
    ws.row_dimensions[row].height = 20
    row += 1

    open_wallets = sorted(
        [wallet for wallet in wallets if wallet.balance != 0],
        key=lambda wallet: (
            user_label(wallet.user_id).lower(),
            wallet.currency_id,
        ),
    )
    if not open_wallets:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws[f"B{row}"] = "No non-zero wallets"
        style(f"B{row}", size=10, color=C["gray"], horizontal="center", border=all_borders())
        ws.row_dimensions[row].height = 20
        row += 1
    else:
        last_owner = None
        for index, wallet in enumerate(open_wallets):
            user = user_map.get(wallet.user_id)
            currency = currency_map.get(wallet.currency_id)
            owner = user_label(wallet.user_id)
            owner_key = owner.lower()
            show_owner = owner_key != last_owner
            fill = C["lightBg"] if index % 2 == 1 else C["white"]
            values = {
                "B": owner if show_owner else "",
                "C": user.role.value if show_owner and user else "",
                "D": currency.name if currency else wallet.currency_id,
                "E": float(wallet.balance),
            }
            for col, value in values.items():
                ws[f"{col}{row}"] = value
                style(
                    f"{col}{row}",
                    size=9,
                    fill=fill,
                    border=all_borders(),
                    horizontal="right" if col == "E" else "center" if col == "C" else "left",
                )
            ws[f"E{row}"].number_format = money_number_format(currency)
            ws.row_dimensions[row].height = 18
            last_owner = owner_key
            row += 1

    row += 1
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws[f"B{row}"] = (
        f"This report was generated on {generated_at} and includes all FX orders, "
        "house exchanges, and transfers recorded for the selected period, including voided records."
    )
    style(f"B{row}", size=8, color=C["gray"], horizontal="center", wrap_text=True)
    ws.row_dimensions[row].height = 28

    output = BytesIO()
    wb.save(output)
    if from_date or to_date:
        filename_period = f"{from_date.isoformat() if from_date else 'start'}_to_{to_date.isoformat() if to_date else 'today'}"
    else:
        filename_period = "all_time"
    return output.getvalue(), f"full_activity_report_{filename_period}.xlsx"


def build_full_activity_report_pdf(
    *,
    users: list[User],
    currencies: list[Currency],
    wallets: list[Wallet],
    orders: list[Order],
    house_exchanges: list[HouseExchange],
    expenses: list[Expense],
    journals: list[JournalEntry],
    wallet_adjustments: list[WalletAdjustment],
    from_date: date | None,
    to_date: date | None,
) -> tuple[bytes, str]:
    generated_at = fmt_report_datetime(datetime.now(UTC))
    period_label = _format_period_label(from_date, to_date)
    user_map = {user.id: user for user in users}
    wallet_map = {wallet.id: wallet for wallet in wallets}
    currency_map = {currency.ticker: currency for currency in currencies}

    def user_label(user_id: int | None) -> str:
        if user_id is None:
            return ""
        user = user_map.get(user_id)
        if not user:
            return str(user_id)
        return user_full_name(user)

    def note_with_void(note: str | None, void_reason: str | None) -> str:
        parts = []
        if note:
            parts.append(note)
        if void_reason:
            parts.append(f"Void reason: {void_reason}")
        return " | ".join(parts)

    def wallet_user_label(wallet_id: int) -> str:
        wallet = wallet_map.get(wallet_id)
        if not wallet:
            return f"Wallet #{wallet_id}"
        return user_label(wallet.user_id)

    tx_rows: list[dict[str, object]] = []
    for order in orders:
        in_curr = currency_map.get(order.currency_in_id)
        out_curr = currency_map.get(order.currency_out_id)
        tx_rows.append(
            {
                "date": order.created_at,
                "dateLabel": fmt_report_datetime(order.created_at),
                "category": "FX Order",
                "party": user_label(order.client_id),
                "type": order.order_type.value,
                "description": (
                    f"{in_curr.name if in_curr else order.currency_in_id} → "
                    f"{out_curr.name if out_curr else order.currency_out_id} "
                    f"@ {float(order.exchange_rate):.4f}"
                ),
                "sent": f"{fmt_currency_money(order.amount_in, in_curr)} {order.currency_in_id}",
                "received": f"{fmt_currency_money(order.amount_out, out_curr)} {order.currency_out_id}",
                "note": note_with_void(order.description, order.void_reason),
                "status": "Voided" if order.voided_at else "Active",
            }
        )

    for exchange in house_exchanges:
        from_curr = currency_map.get(exchange.currency_from_id)
        to_curr = currency_map.get(exchange.currency_to_id)
        tx_rows.append(
            {
                "date": exchange.created_at,
                "dateLabel": fmt_report_datetime(exchange.created_at),
                "category": "House Exchange",
                "party": user_label(exchange.house_id),
                "type": "Exchange",
                "description": (
                    f"{from_curr.name if from_curr else exchange.currency_from_id} → "
                    f"{to_curr.name if to_curr else exchange.currency_to_id} "
                    f"@ {float(exchange.exchange_rate):.4f}"
                ),
                "sent": f"{fmt_currency_money(exchange.amount_from, from_curr)} {exchange.currency_from_id}",
                "received": f"{fmt_currency_money(exchange.amount_to, to_curr)} {exchange.currency_to_id}",
                "note": note_with_void(exchange.description, exchange.void_reason),
                "status": "Voided" if exchange.voided_at else "Active",
            }
        )

    for expense in expenses:
        currency = currency_map.get(expense.currency_id)
        kind_label = "Withdrawal" if expense.expense_type == ExpenseType.WITHDRAWAL else "Expense"
        party = user_label(expense.house_id)
        if expense.recipient_user_id is not None:
            party = f"{party} → {user_label(expense.recipient_user_id)}"
        tx_rows.append(
            {
                "date": expense.created_at,
                "dateLabel": fmt_report_datetime(expense.created_at),
                "category": kind_label,
                "party": party,
                "type": expense.currency_id,
                "description": f"{currency.name if currency else expense.currency_id} {kind_label.lower()}",
                "sent": f"{fmt_currency_money(expense.amount, currency)} {expense.currency_id}",
                "received": "",
                "note": note_with_void(expense.description, expense.void_reason),
                "status": "Voided" if expense.voided_at else "Active",
            }
        )

    for entry in journals:
        currency = currency_map.get(entry.currency_id)
        tx_rows.append(
            {
                "date": entry.created_at,
                "dateLabel": fmt_report_datetime(entry.created_at),
                "category": "Transfer",
                "party": f"{wallet_user_label(entry.from_wallet_id)} → {wallet_user_label(entry.to_wallet_id)}",
                "type": entry.currency_id,
                "description": f"{currency.name if currency else entry.currency_id} transfer",
                "sent": f"{fmt_currency_money(entry.amount, currency)} {entry.currency_id}",
                "received": f"{fmt_currency_money(entry.amount, currency)} {entry.currency_id}",
                "note": note_with_void(entry.description, entry.void_reason),
                "status": "Voided" if entry.voided_at else "Active",
            }
        )

    for adjustment in wallet_adjustments:
        wallet = wallet_map.get(adjustment.wallet_id)
        currency = currency_map.get(adjustment.currency_id)
        amount = adjustment.amount_delta
        tx_rows.append(
            {
                "date": adjustment.created_at,
                "dateLabel": fmt_report_datetime(adjustment.created_at),
                "category": "Wallet Adjustment",
                "party": user_label(wallet.user_id) if wallet else f"Wallet #{adjustment.wallet_id}",
                "type": adjustment.currency_id,
                "description": f"{currency.name if currency else adjustment.currency_id} balance adjustment",
                "sent": f"{fmt_currency_money(abs(amount), currency)} {adjustment.currency_id}" if amount < 0 else "",
                "received": f"{fmt_currency_money(amount, currency)} {adjustment.currency_id}" if amount > 0 else "",
                "note": adjustment.reason,
                "status": "Active",
            }
        )
    tx_rows.sort(key=lambda item: item["date"])

    styles = pdf_styles()
    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )
    story = [
        pdf_paragraph("FULL ACTIVITY REPORT", styles["title"]),
        pdf_paragraph("FX Orders & Transfers", styles["subtitle"]),
        pdf_paragraph(f"Report Period: {period_label}    Generated: {generated_at}", styles["small"]),
        Spacer(1, 6),
        pdf_paragraph(f"TRANSACTION HISTORY ({len(tx_rows)} transactions)", styles["section"]),
    ]
    table_rows = [
        [
            row["dateLabel"],
            row["category"],
            row["party"],
            row["type"],
            row["description"],
            row["sent"],
            row["received"],
            row["note"],
            row["status"],
        ]
        for row in tx_rows
    ] or [["No transactions in this period", "", "", "", "", "", "", "", ""]]
    story.append(
        pdf_table(
            ["Date & Time", "Category", "Party", "Type", "Description", "Sent", "Received", "Note", "Status"],
            table_rows,
            [24 * mm, 24 * mm, 37 * mm, 17 * mm, 48 * mm, 29 * mm, 31 * mm, 49 * mm, 18 * mm],
            right_columns={5, 6},
        )
    )

    open_wallets = sorted(
        [wallet for wallet in wallets if wallet.balance != 0],
        key=lambda wallet: (
            user_label(wallet.user_id).lower(),
            wallet.currency_id,
        ),
    )
    wallet_rows: list[list[object]] = []
    last_owner = None
    for wallet in open_wallets:
        user = user_map.get(wallet.user_id)
        currency = currency_map.get(wallet.currency_id)
        owner = user_label(wallet.user_id)
        owner_key = owner.lower()
        show_owner = owner_key != last_owner
        wallet_rows.append(
            [
                owner if show_owner else "",
                user.role.value if show_owner and user else "",
                currency.name if currency else wallet.currency_id,
                fmt_currency_money(wallet.balance, currency),
            ]
        )
        last_owner = owner_key
    if not wallet_rows:
        wallet_rows.append(["No non-zero wallets", "", "", ""])

    story.append(Spacer(1, 8))
    story.append(pdf_paragraph("NON-ZERO WALLETS", styles["section"]))
    story.append(
        pdf_table(
            ["Name", "Role", "Currency", "Balance"],
            wallet_rows,
            [70 * mm, 30 * mm, 70 * mm, 40 * mm],
            right_columns={3},
        )
    )
    story.append(Spacer(1, 6))
    story.append(
        pdf_paragraph(
            f"This report was generated on {generated_at} and includes all FX orders, house exchanges, and transfers recorded for the selected period, including voided records.",
            styles["small"],
        )
    )
    doc.build(story)
    if from_date or to_date:
        filename_period = f"{from_date.isoformat() if from_date else 'start'}_to_{to_date.isoformat() if to_date else 'today'}"
    else:
        filename_period = "all_time"
    return output.getvalue(), f"full_activity_report_{filename_period}.pdf"
